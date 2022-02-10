import openpyxl
import heapq
loc = 'LanguageExchange.xlsx'
wb = openpyxl.load_workbook(loc)
sheet = wb.active
# self.matched = set()
# self.data = {}

matches = []

def main(sheet):
	
	# seek = 'K%s' % (i + 2)   # col K is now the seeking to in excel, if excel changes format this might needs changing
	# offer = 'N%s' % (j + 2)  # same as above, N is for offering to
	seek = 10 # col K is index 10
	offer = 13 # col N is index 13
	for i, row in enumerate(sheet.iter_rows(values_only = True)):
		for j, row2 in enumerate(sheet.iter_rows(values_only = True)):
			
			if j == i:
				continue
			#if j in self.matched:
			#	continue
			if row2[offer] == row[seek] and row2[seek] == row[offer]:
				
				isMatch(i+1, j+1) # row numbers i+1 and j+1 is a match
	printMatch()

def isMatch(row1, row2): # print format: num1, num1 seeking, num1 offering, num2, num2 seeking, num2 offering
	heapq.heappush(matches, (row1, sheet['K%s'%(row1)].value, sheet['N%s'%(row1)].value, row2, sheet['K%s'%(row2)].value, sheet['N%s'%(row2)].value))

def printMatch():
	
	while matches:
		print(heapq.heappop(matches))


if __name__ == '__main__':
	print('student1','seeking','offering','student2', 'seeking', 'offering')
	main(sheet)